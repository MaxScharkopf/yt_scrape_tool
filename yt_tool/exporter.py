"""Export database contents to CSV."""

import csv
import logging
import sqlite3
from datetime import datetime
from pathlib import Path

from .config import DB_FILE, EXPORT_DIR

logger = logging.getLogger(__name__)


def cmd_export(query: str | None = None) -> None:
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if query:
        c.execute(
            "SELECT title, channel, duration, views, views_int, url, query, scraped_at "
            "FROM videos WHERE query = ? ORDER BY scraped_at DESC",
            (query,),
        )
        filename = f"youtube_{query.replace(' ', '_')}_{timestamp}.csv"
    else:
        c.execute(
            "SELECT title, channel, duration, views, views_int, url, query, scraped_at "
            "FROM videos ORDER BY scraped_at DESC"
        )
        filename = f"youtube_all_{timestamp}.csv"

    rows = c.fetchall()
    conn.close()

    if not rows:
        print("\n  No data found to export.\n")
        return

    out_path = Path(EXPORT_DIR) / filename
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Title", "Channel", "Duration", "Views", "Views (int)", "URL", "Query", "Scraped At"])
        writer.writerows(rows)

    print(f"\n✅ Exported {len(rows)} videos to: {out_path}")
    print("   Open it in Excel or Google Sheets!\n")
    logger.info("Exported %d row(s) to %s.", len(rows), out_path)


def cmd_export_excel(query: str | None = None) -> None:
    """Export database to a formatted .xlsx file using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n  openpyxl not installed. Run: pip install openpyxl\n")
        logger.error("openpyxl is not installed.")
        return

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if query:
        c.execute(
            "SELECT title, channel, duration, views, views_int, url, query, scraped_at "
            "FROM videos WHERE query = ? ORDER BY scraped_at DESC",
            (query,),
        )
        filename = f"youtube_{query.replace(' ', '_')}_{timestamp}.xlsx"
    else:
        c.execute(
            "SELECT title, channel, duration, views, views_int, url, query, scraped_at "
            "FROM videos ORDER BY scraped_at DESC"
        )
        filename = f"youtube_all_{timestamp}.xlsx"

    rows = c.fetchall()
    conn.close()

    if not rows:
        print("\n  No data found to export.\n")
        return

    headers = ["Title", "Channel", "Duration", "Views", "Views (int)", "URL", "Query", "Scraped At"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YouTube Data"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    out_path = Path(EXPORT_DIR) / filename
    wb.save(out_path)
    print(f"\n✅ Exported {len(rows)} videos to: {out_path}\n")
    logger.info("Excel export: %d row(s) to %s.", len(rows), out_path)
